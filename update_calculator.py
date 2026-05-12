#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
update_calculator.py — обновляет calculator.html из prices.xlsx

Использование:
    python3 update_calculator.py             # обновить один раз
    python3 update_calculator.py --watch     # следить за prices.xlsx и пересобирать при изменении

Зависимости:
    pip3 install openpyxl
"""

import json
import sys
import os
import time
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("✗ Не установлен openpyxl. Установите командой:")
    print("    pip3 install openpyxl")
    sys.exit(1)

# Файлы — все рядом со скриптом
HERE = Path(__file__).parent
PRICES_FILE   = HERE / 'prices.xlsx'
TEMPLATE_FILE = HERE / 'calculator-template.html'
OUTPUT_FILE   = HERE / 'calculator.html'


# ============================================================
# ЧТЕНИЕ EXCEL
# ============================================================
def read_prices(xlsx_path: Path) -> dict:
    """Парсит prices.xlsx и возвращает dict для инжекта в HTML."""
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Не найден файл {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)

    # ----- Лист «Опции» -----
    if 'Опции' not in wb.sheetnames:
        raise ValueError("В prices.xlsx нет листа «Опции»")
    ws = wb['Опции']

    foundation, walls, roof, facade = [], [], [], []
    category_map = {
        'Фундамент': foundation,
        'Стены':     walls,
        'Кровля':    roof,
        'Фасад':     facade,
    }

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        category, key, name, price, active = (row + (None,) * 5)[:5]
        if category not in category_map:
            continue
        try:
            active_flag = int(active) if active is not None else 1
        except (ValueError, TypeError):
            active_flag = 1
        if active_flag != 1:
            continue
        if not all([key, name, price is not None]):
            continue
        category_map[category].append({
            'key':   str(key).strip(),
            'name':  str(name).strip(),
            'price': float(price),
        })

    # ----- Лист «Коэффициенты» -----
    if 'Коэффициенты' not in wb.sheetnames:
        raise ValueError("В prices.xlsx нет листа «Коэффициенты»")
    ws2 = wb['Коэффициенты']

    floors = []
    params = {}
    section = None
    for row in ws2.iter_rows(values_only=True):
        if not row or all(c is None for c in row):
            continue
        first = str(row[0]).strip() if row[0] else ''
        if first.startswith('ЭТАЖНОСТЬ'):
            section = 'floors'; continue
        if first.startswith('ОБЩИЕ'):
            section = 'params'; continue
        if first in ('Ключ',):
            continue  # шапка таблицы

        if section == 'floors':
            key, name, coef = (row + (None,) * 3)[:3]
            if key is not None and coef is not None:
                floors.append({
                    'key':  str(key).strip(),
                    'name': str(name).strip() if name else str(key),
                    'coef': float(coef),
                })
        elif section == 'params':
            key, _desc, value = (row + (None,) * 3)[:3]
            if key is not None and value is not None:
                params[str(key).strip()] = (
                    float(value) if isinstance(value, (int, float)) else value
                )

    # ----- Лист «Контакты» -----
    contacts = {}
    title = subtitle = pos_button = ''
    if 'Контакты' in wb.sheetnames:
        ws3 = wb['Контакты']
        contact_map = {
            'Название компании':         'company',
            'Телефон (отображение)':     'phone_display',
            'Телефон (для tel: ссылки)': 'phone_tel',
            'WhatsApp (номер)':          'whatsapp',
            'Заголовок калькулятора':    '_title',
            'Подзаголовок':              '_subtitle',
            'Текст кнопки ПОС':          '_pos_button',
        }
        for row in ws3.iter_rows(min_row=3, values_only=True):
            if not row or row[0] is None:
                continue
            field, value = (row + (None, None))[:2]
            key = contact_map.get(str(field).strip())
            if key and value is not None:
                val_str = str(value).strip()
                if key == '_title':       title = val_str
                elif key == '_subtitle':  subtitle = val_str
                elif key == '_pos_button': pos_button = val_str
                else:                     contacts[key] = val_str

    return {
        'foundation': foundation,
        'walls':      walls,
        'roof':       roof,
        'facade':     facade,
        'floors':     floors,
        'params':     params,
        'contacts':   contacts,
        '_meta': {
            'title':      title or 'Калькулятор стоимости строительства',
            'subtitle':   subtitle or 'Рассчитайте дом под ключ за 30 секунд',
            'pos_button': pos_button or 'Проект организации строительства',
        },
    }


# ============================================================
# ВАЛИДАЦИЯ
# ============================================================
def validate(data: dict) -> list:
    errors = []
    for cat in ('foundation', 'walls', 'roof', 'facade'):
        if not data[cat]:
            errors.append(f"Нет ни одной активной опции в категории «{cat}»")
    if not data['floors']:
        errors.append("Нет данных по этажности")
    required_params = ['design_markup', 'pos_fixed', 'floor_height',
                       'facade_coef', 'area_min', 'area_max', 'area_default']
    for p in required_params:
        if p not in data['params']:
            errors.append(f"В «Коэффициентах» отсутствует параметр {p}")
    return errors


# ============================================================
# ГЕНЕРАЦИЯ HTML
# ============================================================
def generate_html(data: dict) -> str:
    template = TEMPLATE_FILE.read_text(encoding='utf-8')

    # Данные для JS (без _meta — он в плейсхолдерах)
    js_data = {k: v for k, v in data.items() if not k.startswith('_')}
    prices_json = json.dumps(js_data, ensure_ascii=False, indent=2)

    meta = data['_meta']
    contacts = data['contacts']

    # Подстановка плейсхолдеров
    out = template
    out = out.replace('/*{{PRICES_INJECTION}}*/null', prices_json)
    out = out.replace('{{TITLE}}',         meta['title'])
    out = out.replace('{{SUBTITLE}}',      meta['subtitle'])
    out = out.replace('{{POS_BUTTON}}',    meta['pos_button'])
    out = out.replace('{{COMPANY}}',       contacts.get('company', ''))
    out = out.replace('{{PHONE_DISPLAY}}', contacts.get('phone_display', ''))
    out = out.replace('{{PHONE_TEL}}',     contacts.get('phone_tel', ''))
    out = out.replace('{{GENERATED_AT}}',
                      datetime.now().strftime('%d.%m.%Y %H:%M'))
    return out


# ============================================================
# ЦИКЛ ОБНОВЛЕНИЯ
# ============================================================
def update_once() -> bool:
    try:
        data = read_prices(PRICES_FILE)
    except Exception as e:
        print(f"✗ Ошибка чтения {PRICES_FILE.name}: {e}")
        return False

    errors = validate(data)
    if errors:
        print("✗ Найдены ошибки в prices.xlsx:")
        for e in errors:
            print(f"   – {e}")
        return False

    html = generate_html(data)
    OUTPUT_FILE.write_text(html, encoding='utf-8')

    total_options = sum(len(data[c]) for c in ('foundation', 'walls', 'roof', 'facade'))
    print(f"✓ {datetime.now():%H:%M:%S} → {OUTPUT_FILE.name} "
          f"({total_options} опций, "
          f"{len(data['foundation'])} фунд., "
          f"{len(data['walls'])} стен, "
          f"{len(data['roof'])} кровли, "
          f"{len(data['facade'])} фасад)")
    return True


def watch():
    """Следит за изменением prices.xlsx и автоматически пересобирает."""
    print(f"👀 Слежу за {PRICES_FILE.name}. Ctrl+C — выход.")
    last_mtime = 0
    while True:
        try:
            mtime = PRICES_FILE.stat().st_mtime
            if mtime != last_mtime:
                if last_mtime != 0:
                    print("  ↻ Обнаружено изменение, обновляю…")
                update_once()
                last_mtime = mtime
            time.sleep(1)
        except KeyboardInterrupt:
            print("\n👋 Остановлено.")
            break
        except FileNotFoundError:
            print(f"✗ Файл {PRICES_FILE} пропал. Жду…")
            time.sleep(2)


# ============================================================
# MAIN
# ============================================================
if __name__ == '__main__':
    if '--watch' in sys.argv or '-w' in sys.argv:
        update_once()
        watch()
    else:
        ok = update_once()
        sys.exit(0 if ok else 1)
